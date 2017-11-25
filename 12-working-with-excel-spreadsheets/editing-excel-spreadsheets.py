import openpyxl
import os

wb = openpyxl.Workbook() # creates and stores new spreadsheet / workbook

# print(wb.get_sheet_names())
# sheet = wb.get_sheet_by_name('Sheet')

# sheet['A1'].value == None # returns true
# sheet['A1'].value = 42
# sheet['A2'].value = 'Hello'
#
# sheet2 = wb.create_sheet()
# sheet2.title = 'My New Sheet Name'
#
# wb.save('editing-example.xlsx')

wb.create_sheet(index=0, title='My Other Sheet')
wb.save('editing-example2.xlsx')
