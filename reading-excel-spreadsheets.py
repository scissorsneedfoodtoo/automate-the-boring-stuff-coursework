import openpyxl
import os

# os.chdir('...') # change dir to one with excel file if needed

workbook = openpyxl.load_workbook('example.xlsx')
sheet = workbook.get_sheet_by_name('Sheet1')
# workbook.get_sheet_names() # returns list of sheet names

cell = sheet['A1']
print(str(cell.value))

print(str(sheet['B1'].value))
print(str(sheet['C1'].value))

# sheet.cell(row=1, column=2)

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
